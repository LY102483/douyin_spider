import datetime
import os
import time

import requests
import openpyxl
from selenium import webdriver
from selenium.webdriver.common.by import By

'''

核心函数start

'''


# 创建文件夹函数
def mkdir(keyword):
    '''
    创建指定的文件夹
    :param path: 文件夹路径，字符串格式
    :return: True(新建成功) or False(文件夹已存在，新建失败)
    '''
    # 判断路径是否存在
    # 存在     True
    # 不存在   False
    path = os.getcwd() + "/videos/" + keyword
    isExists = os.path.exists(path)

    # 判断结果
    if not isExists:
        # 如果不存在则创建目录
        # 创建目录操作函数
        os.makedirs(path)
        return True
    else:
        # 如果目录存在则不创建
        return False


# 检查excel表是否存在
def createExcel():
    # 判断路径是否存在
    # 存在     True
    # 不存在   False
    path = os.getcwd() + "/data.xlsx"
    isExists = os.path.exists(path)
    # 判断结果
    if isExists:
        return True
    else:
        return False


# 下载函数
def downloadVideo(videoUrl, keyword, rank, nowTime):
    keywordOld = keyword
    keyword = nowTime + keyword
    mkdir(keyword)
    video_content = requests.get(url=videoUrl).content
    with open(os.getcwd() + "/videos/" + keyword + "/" + rank + '.mp4', mode='wb') as f:
        f.write(video_content)
    print('视频保存成功 关键字:', keywordOld, '排名:', rank, '下载地址:', videoUrl)


def readExcel():
    # 打开excel文件
    wb = openpyxl.load_workbook('data.xlsx')
    # 获取活跃表对象
    sheet = wb.active
    # 获取第二列的所有内容
    row_num = sheet.max_row  # 获取当前表中最大的行数
    for row in range(1, row_num + 1):
        cell = sheet.cell(row, 1)
        keywords.append(cell.value)


# 通过关键字获取视频链接
def videoSrc(keyword, nowTime):
    # 重试方法
    attempts = 0
    success = False
    while attempts < 6 and not success:
        try:
            searchUrl = "https://www.douyin.com/search/" + keyword + "?publish_time=1&sort_type=1&source=tab_search&type=general"
            # url = searchUrl
            driver.get(searchUrl)
            time.sleep(2)
            # 滑动到底部
            # ActionChains(driver).send_keys(Keys.END).perform()

            # 开始调用函数正式开始获取
            for x in range(1, 6):
                # 排名前五的视频
                rank = str(x)
                # source src定位
                source = '/html/body/div[1]/div/div[2]/div/div[3]/div[1]/ul/li[' + rank + ']/div/div/div[3]/div/div/div[1]/div[1]/div/div[2]/div[1]/video/source[1]'
                # 播放按钮定位
                playButton = '/html/body/div[1]/div/div[2]/div/div[3]/div[1]/ul/li[' + rank + ']/div/div/div[3]/div/div/div[1]/div[1]/div/div[2]'

                # 滑动窗口到元素位置
                eles = driver.find_elements_by_xpath(playButton)
                ele = eles[0]
                driver.execute_script('arguments[0].scrollIntoView()', ele)
                clickCnt = 0  # 记录获取播放地址的次数
                videoUrlBool = False  # 记录是否成功获取播放地址
                while clickCnt < 6 and not videoUrlBool:
                    try:
                        # 点击播放按钮
                        driver.find_element(By.XPATH, playButton).click()
                        time.sleep(2)
                        # 获取视频地址
                        videoUrl = driver.find_element(By.XPATH, source).get_attribute('src')
                        downloadVideo(videoUrl, keyword, rank, nowTime)
                        videoUrlBool = True
                    except Exception :
                        print("未获取到链接！")
                        clickCnt += 1

            success = True
            print("关键字： " + keyword + " 数据已全部获取\n")
        except Exception as errorInfo:
            attempts += 1
            if (attempts < 6):
                print("发生异常，异常原因：" + str(errorInfo))
                print("关键字:", keyword, " 正在进行第", attempts, "次重试\n")
            else:
                errorKeywords.append(keyword)


'''

核心函数end

'''
# 模拟浏览器代理
headers = {
    'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; WOW64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/82.0.4051.0 Safari/537.36 Edg/82.0.425.0'}

# 计时器
start_time = time.time()
# 设置Chrome浏览器
chrome_options = webdriver.ChromeOptions()
# 隐身访问
# chrome_options.add_argument('--incognito')
# 不加载图片, 提升速度
# chrome_options.add_argument('--blink-settings=imagesEnabled=false')
# 不打开浏览器窗口
# chrome_options.add_argument("headless")

if not createExcel():
    print("未找到excel表,请在当前路径下创建data.xlsx,并将关键字填写至第一列！！")
else:
    # 定义一个列表用来存储没有获取成功的数据
    errorKeywords = []
    # 定义一个列表用来存储关键字
    keywords = []
    readExcel()
    try:
        print("请在新打开的网页中进行登陆操作！\n")
        time.sleep(3)
        # 告诉编译器chromedriver在哪个位置并注册
        driver = webdriver.Chrome(os.getcwd() + "/chromedriver.exe", chrome_options=chrome_options)
        # 登陆操作(20s时间)
        driver.get("https://www.douyin.com")
        time.sleep(20)
        # 获取当前任务执行时间
        nowTime = datetime.datetime.now()
        nowTime = nowTime.strftime("%Y%m%d%H%M")

        for keyword in keywords:
            videoSrc(keyword, nowTime)

        driver.close()


    except Exception as errorInfo:
        print("结束原因：" + str(errorInfo))
    finally:
        print("\n程序运行结束")
        if (len(errorKeywords) == 0):
            print("关键字全部成功获取")
        else:
            print("以下关键字未能成功获取成功获取")
            print(errorKeywords)
